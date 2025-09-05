import io, os
import arabic_reshaper
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from bidi.algorithm import get_display
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from django.conf import settings
from django.urls import reverse_lazy
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.http import FileResponse, HttpResponse
from django.db import IntegrityError, DatabaseError
from django.db.models import Q

from employees.mixins import PermissionMixin, ExportMixin
from .models import Device, DeviceType, Custody, Status
from .forms import DeviceForm, DeviceAccessoryFormSet, CustodyForm, DeviceCustodyFormSet

# Device Types (manage manually)
@login_required
@permission_required("devices.view_devicetype", raise_exception=True)
def devices_type_manage(request):
    edit_id = request.GET.get("edit")
    delete_id = request.GET.get("delete")

    # --- Delete ---
    if delete_id:
        if not request.user.has_perm("devices.delete_devicetype") and not request.user.is_superuser:
            messages.error(request, _("You do not have permission to delete a device type"))
            return redirect("devicestype_manage")

        obj = get_object_or_404(DeviceType, pk=delete_id)
        try:
            obj.delete()
            messages.success(request, _("Device type deleted successfully"))
        except Exception:
            messages.error(request, _("An unexpected error occurred while deleting"))
        return redirect("devicestype_manage")

    # --- Add / Edit ---
    if request.method == "POST":
        name_ar = request.POST.get("name_ar", "").strip()
        name_en = request.POST.get("name_en", "").strip()
        if not name_en or not name_ar:
            messages.error(request, _("Both English and Arabic names are required."))
            return redirect("supplycategory_manage")

        try:
            if edit_id:  # Edit
                if not request.user.has_perm("devices.change_devicetype") and not request.user.is_superuser:
                    messages.error(request, _("You do not have permission to edit a device type"))
                    return redirect("devicestype_manage")

                obj = get_object_or_404(DeviceType, pk=edit_id)

                # Duplicate checks
                if DeviceType.objects.exclude(pk=obj.pk).filter(name_en=name_en).exists():
                    messages.error(request, _("A device type with this English name already exists"))
                    return redirect("devicestype_manage")
                if DeviceType.objects.exclude(pk=obj.pk).filter(name_ar=name_ar).exists():
                    messages.error(request, _("A device type with this Arabic name already exists"))
                    return redirect("devicestype_manage")

                obj.name_ar = name_ar
                obj.name_en = name_en
                obj.save()
                messages.success(request, _("Device type updated successfully"))

            else:  # Add
                if not request.user.has_perm("devices.add_devicetype") and not request.user.is_superuser:
                    messages.error(request, _("You do not have permission to add a device type"))
                    return redirect("devicestype_manage")

                # Duplicate checks
                if DeviceType.objects.filter(name_en=name_en).exists():
                    messages.error(request, _("A device type with this English name already exists"))
                    return redirect("devicestype_manage")
                if DeviceType.objects.filter(name_ar=name_ar).exists():
                    messages.error(request, _("A device type with this Arabic name already exists"))
                    return redirect("devicestype_manage")

                DeviceType.objects.create(name_ar=name_ar, name_en=name_en)
                messages.success(request, _("Device type added successfully"))

        except IntegrityError:
            messages.error(request, _("A database integrity error occurred. Please check your data."))
        except DatabaseError:
            messages.error(request, _("A database error occurred. Please try again later."))
        except Exception:
            messages.error(request, _("An unexpected error occurred. Please try again."))

        return redirect("devicestype_manage")

    # --- Render ---
    types = DeviceType.objects.all()
    current_edit = None
    if edit_id:
        current_edit = get_object_or_404(DeviceType, pk=edit_id)

    return render(request, "devices/devicetype_manage.html", {
        "device_types": types,
        "current_edit": current_edit,
    })


# Devices
class DeviceListView(PermissionMixin, ListView):
    model = Device
    template_name = "devices/list.html"
    context_object_name = "devices"
    permission_required = "devices.view_device"
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related("device_type")
        search_query = self.request.GET.get("q", "").strip()
        device_type_id = self.request.GET.get("device_type", "").strip()
        status = self.request.GET.get("status", "").strip()

        if search_query:
            queryset = queryset.filter(
                Q(name_en__icontains=search_query) |
                Q(name_ar__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(brand__icontains=search_query)
            )

        if device_type_id.isdigit():
            queryset = queryset.filter(device_type_id=device_type_id)

        if status in dict(Status.choices):
            queryset = queryset.filter(status=status)

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["device_types"] = DeviceType.objects.all()
        context["statuses"] = dict(Status.choices)
        context["selected_type"] = self.request.GET.get("device_type", "").strip()
        context["selected_status"] = self.request.GET.get("status", "").strip()
        context["search_query"] = self.request.GET.get("q", "").strip()
        return context


class DeviceDetailView(PermissionMixin, DetailView):
    model = Device
    template_name = "devices/detail.html"
    context_object_name = "device"
    permission_required = "devices.view_device"


class DeviceCreateView(PermissionMixin, CreateView):
    model = Device
    form_class = DeviceForm
    template_name = "devices/form.html"
    success_url = reverse_lazy("device_list")
    permission_required = "devices.add_device"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["accessory_formset"] = DeviceAccessoryFormSet(self.request.POST)
        else:
            context["accessory_formset"] = DeviceAccessoryFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        accessory_formset = context["accessory_formset"]

        if accessory_formset.is_valid():
            self.object = form.save()  # save device first
            accessory_formset.instance = self.object  # link accessories to device
            accessory_formset.save()
            messages.success(self.request, _("Device created successfully"))
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class DeviceUpdateView(PermissionMixin, UpdateView):
    model = Device
    form_class = DeviceForm
    template_name = "devices/form.html"
    success_url = reverse_lazy("device_list")
    permission_required = "devices.change_device"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["accessory_formset"] = DeviceAccessoryFormSet(
                self.request.POST, instance=self.object
            )
        else:
            context["accessory_formset"] = DeviceAccessoryFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        accessory_formset = context["accessory_formset"]

        if accessory_formset.is_valid():
            self.object = form.save()
            accessory_formset.instance = self.object
            accessory_formset.save()
            messages.success(self.request, _("Device updated successfully"))
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class DeviceDeleteView(DeleteView):
    model = Device
    template_name = "devices/confirm_delete.html"
    success_url = reverse_lazy("device_list")
    permission_required = "devices.delete_device"

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, _("Device deleted successfully"))
        return super().delete(request, *args, **kwargs)


# Custody Views
class CustodyListView(PermissionMixin, ListView):
    model = Custody
    template_name = "custody/list.html"
    context_object_name = "custodies"
    permission_required = "devices.view_custody"
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related("employee").prefetch_related("devices__device")
        
        # Restrict access for non-managers
        if not self.request.user.groups.filter(name__in=["Manager", "It Employee"]).exists() and not self.request.user.is_superuser:
            queryset = queryset.filter(employee=self.request.user)

        search_query = self.request.GET.get("q", "").strip()
        device_type_id = self.request.GET.get("device_type", "").strip()
        is_returned = self.request.GET.get("is_returned")

        # Search across employee profile and device fields
        if search_query:
            queryset = queryset.filter(
                Q(employee__user__username__icontains=search_query) |
                Q(employee__full_name_en__icontains=search_query) |
                Q(employee__full_name_ar__icontains=search_query) |
                Q(devices__device__name_en__icontains=search_query) |
                Q(devices__device__name_ar__icontains=search_query) |
                Q(devices__device__serial_number__icontains=search_query)
            ).distinct()

        # Filter by device type
        if device_type_id.isdigit():
            queryset = queryset.filter(devices__device__device_type_id=device_type_id)

        # Filter by returned status
        if is_returned == "yes":
            queryset = queryset.filter(return_date__isnull=False)
        elif is_returned == "no":
            queryset = queryset.filter(return_date__isnull=True)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["device_types"] = DeviceType.objects.all()
        context["search_query"] = self.request.GET.get("q", "").strip()
        context["selected_type"] = self.request.GET.get("device_type", "").strip()
        context["selected_returned"] = self.request.GET.get("is_returned", "").strip()
        return context


class CustodyDetailView(PermissionMixin, DetailView):
    model = Custody
    template_name = "custody/detail.html"
    context_object_name = "custody"
    permission_required = "devices.view_custody"


class CustodyCreateView(PermissionMixin, CreateView):
    model = Custody
    form_class = CustodyForm
    template_name = "custody/form.html"
    success_url = reverse_lazy("custody_list")
    permission_required = "devices.add_custody"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["devicecustody_formset"] = DeviceCustodyFormSet(self.request.POST)
        else:
            context["devicecustody_formset"] = DeviceCustodyFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context["devicecustody_formset"]
        if form.is_valid() and formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            messages.success(self.request, _("Custody created successfully."))
            return redirect(self.success_url)
        messages.error(self.request, _("There was an error creating the custody."))
        return self.render_to_response(self.get_context_data(form=form))


class CustodyUpdateView(PermissionMixin, UpdateView):
    model = Custody
    form_class = CustodyForm
    template_name = "custody/form.html"
    success_url = reverse_lazy("custody_list")
    permission_required = "devices.change_custody"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["devicecustody_formset"] = DeviceCustodyFormSet(
                self.request.POST, instance=self.object
            )
        else:
            context["devicecustody_formset"] = DeviceCustodyFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context["devicecustody_formset"]
        if form.is_valid() and formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            messages.success(self.request, _("Custody updated successfully."))
            return redirect(self.success_url)
        messages.error(self.request, _("There was an error updating the custody."))
        return self.render_to_response(self.get_context_data(form=form))


class CustodyDeleteView(PermissionMixin, DeleteView):
    model = Custody
    template_name = "custody/delete.html"
    success_url = reverse_lazy("custody_list")
    permission_required = "devices.delete_custody"

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, _("Custody deleted successfully"))
        return super().delete(request, *args, **kwargs)


class CustodyExportView(PermissionMixin, ExportMixin):
    model = Custody
    filename_prefix = "custody"
    action = None
    permission_required = "devices.change_custody"

    def dispatch(self, request, *args, **kwargs):
        # get action from url (pdf or excel)
        self.action = kwargs.pop("action", None)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        if self.action == "pdf":
            return self.export_pdf(request, *args, **kwargs)
        elif self.action == "excel":
            return self.export_excel(request, *args, **kwargs)
        else:
            return HttpResponse("Invalid export action", status=400)

    def get_pdf_content(self, obj):
        fields = [
            ("ID", obj.id),
            ("Employee", f"{obj.employee.full_name_en} / {obj.employee.full_name_ar}"),
            ("Department", obj.employee.department),
            ("Job Title", obj.employee.job_title),
            ("Custody Date", obj.custody_date),
            ("Return Date", obj.return_date or "Not returned"),
            ("Notes", obj.notes or "-"),
        ]

        devices_lines = []
        for dc in obj.devices.select_related("device").all():
            devices_lines.append(
                f"{dc.device.name_en} / {dc.device.name_ar} "
                f"({dc.device.serial_number}) - {dc.device.brand} - {dc.device.get_status_display()}"
            )
            if dc.notes:
                devices_lines.append(f"Notes: {dc.notes}")
            for acc in dc.accessories.all():
                devices_lines.append(
                    f"   * {acc.name_en}/{acc.name_ar} ({acc.get_status_display()} - {acc.get_condition_display()})"
                )

        extra_sections = [("Devices", devices_lines)]
        return obj, fields, extra_sections

    def get_excel_content(self, obj):
        sheets = []

        sheet1 = [
            ["Field", "Value"],
            ["ID", obj.id],
            ["Employee", f"{obj.employee.full_name_en} / {obj.employee.full_name_ar}"],
            ["Department", obj.employee.department],
            ["Job Title", obj.employee.job_title],
            ["Custody Date", obj.custody_date],
            ["Return Date", obj.return_date or "Not returned"],
            ["Notes", obj.notes or "-"],
        ]
        sheets.append(("Custody Info", sheet1))

        sheet2 = [["Device", "Serial", "Brand", "Status", "Notes"]]
        for dc in obj.devices.select_related("device").all():
            sheet2.append([
                f"{dc.device.name_en}/{dc.device.name_ar}",
                dc.device.serial_number,
                dc.device.brand,
                dc.device.get_status_display(),
                dc.notes or "",
            ])
        sheets.append(("Devices", sheet2))

        sheet3 = [["Device", "Accessory", "Status", "Condition"]]
        for dc in obj.devices.all():
            for acc in dc.accessories.all():
                sheet3.append([
                    f"{dc.device.name_en}/{dc.device.name_ar}",
                    f"{acc.name_en}/{acc.name_ar}",
                    acc.get_status_display(),
                    acc.get_condition_display(),
                ])
        sheets.append(("Accessories", sheet3))

        return sheets


# =============== PDF Config (Helpers, Fonts, Styles) ===============
class PDFConfig:
    """Centralized configuration for fonts, helpers, and styles."""

    @staticmethod
    def ar_text(txt: str) -> str:
        """Reshape + bidi Arabic text for correct PDF rendering."""
        if not txt:
            return ""
        return get_display(arabic_reshaper.reshape(str(txt)))

    # ---- Fonts ----
    font_dir = os.path.join(settings.BASE_DIR, "static", "font")
    dejavu_path = os.path.join(font_dir, "DejaVuSans.ttf")

    try:
        pdfmetrics.registerFont(TTFont("DejaVu", dejavu_path))
    except Exception:
        pass

    # ---- Styles ----
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Arabic", fontName="DejaVu", alignment=2, fontSize=12, leading=16))
    styles.add(ParagraphStyle(name="English", fontName="DejaVu", alignment=0, fontSize=10))
    styles.add(
        ParagraphStyle(
            name="SectionTitle",
            fontName="DejaVu",
            alignment=1,
            fontSize=13,
            spaceAfter=12,
            spaceBefore=12,
        )
    )


# =============== Header/Footer ===============
class HeaderFooter:
    @staticmethod
    def draw_header(canvas, doc):
        canvas.saveState()
        canvas.setFont("DejaVu", 11)

        # Arabic (right side)
        text_ar = (
            "المملكة العربية السعودية\n"
            "مجلس شؤون الجامعات\n"
            "كليات الأولى الأهلية بالأحساء"
        )
        for i, line in enumerate(text_ar.split("\n")):
            canvas.drawRightString(A4[0] - 30, A4[1] - 30 - (i * 18), PDFConfig.ar_text(line))

        # Logo
        logo_path = os.path.join(settings.BASE_DIR, "static", "img", "logo.png")
        if os.path.exists(logo_path):
            canvas.drawImage(
                logo_path,
                A4[0] / 2 - 30,
                A4[1] - 70,
                width=100,
                height=60,
                preserveAspectRatio=True,
                mask="auto"
            )

        # English (left side)
        left_text = (
            "Kingdom of Saudi Arabia\n"
            "Council Of Universities' Affairs\n"
            "AlOola Colleges Private Al Ahsa"
        )
        for i, line in enumerate(left_text.split("\n")):
            canvas.drawString(30, A4[1] - 30 - (i * 18), line)

        canvas.restoreState()

    @staticmethod
    def draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("DejaVu", 8)
        footer_ar = "الأحساء – المبرز – الحي الأكاديمي، الرمز البريدي: 36429 | Al-Mubarraz, Al-Ahsa - Academic District, Zip Code: 36429"
        footer_en = "Al-Mubarraz, Al-Ahsa - Academic District, Zip Code: 36429"

        # Top line (gold color)
        canvas.setStrokeColorRGB(193 / 255.0, 135 / 255.0, 65 / 255.0)
        canvas.setLineWidth(8)
        canvas.line(0, A4[1], A4[0], A4[1] - 0)

        canvas.drawCentredString(A4[0] / 2, 14, PDFConfig.ar_text(footer_ar))
        canvas.line(0, 0, A4[0], 0)
        canvas.setLineWidth(1)
        canvas.setStrokeColorRGB(0, 0, 0)
        canvas.line(30, 30, A4[0] - 30, 30)
        canvas.restoreState()


# =============== Main Export Class ===============
class CustodyPDFBuilder:
    def __init__(self, obj):
        self.obj = obj
        self.buffer = io.BytesIO()
        self.header_style = ParagraphStyle(
            "HeaderStyle", fontName="DejaVu", fontSize=10, textColor=colors.white, alignment=1
        )

    def build(self):
        doc = BaseDocTemplate(
            self.buffer,
            pagesize=A4,
            leftMargin=30,
            rightMargin=30,
            topMargin=100,
            bottomMargin=40
        )

        frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")

        doc.addPageTemplates([
            PageTemplate(id="main", frames=frame, onPage=HeaderFooter.draw_header, onPageEnd=HeaderFooter.draw_footer)
        ])

        story = []
        story += self._title_section()
        story += self._employee_section()
        story += self._devices_section()
        story += self._declaration_section()
        story += self._signature_section()

        doc.build(story)
        self.buffer.seek(0)
        return self.buffer

    # ---------- Helpers ----------
    def make_paragraph(self, text, style=PDFConfig.styles["Arabic"]):
        """Return a Paragraph with Arabic reshaping + bidi if needed."""
        if text is None:
            text = "-"
        if any("\u0600" <= c <= "\u06FF" for c in text):  # Arabic letters
            reshaped = arabic_reshaper.reshape(str(text))
            bidi_text = get_display(reshaped).replace("\n", "<br/>")
            return Paragraph(bidi_text, style)
        return Paragraph(str(text), style)

    # ---------- Sections ----------
    def _title_section(self):
        return [
            Paragraph(PDFConfig.ar_text("نموذج تسليم عهدة / Custody Receipt Form"), PDFConfig.styles["SectionTitle"]),
            Spacer(1, 20),
        ]

    def _employee_section(self):
        fields = [
            (str(self.obj.custody_date), PDFConfig.ar_text("التاريخ / Date : ")),
            (PDFConfig.ar_text("اسم الموظف / Employee Name :"), ""),
            (f"{PDFConfig.ar_text(self.obj.employee.full_name_ar)} / {self.obj.employee.full_name_en}", ""),
            (PDFConfig.ar_text(self.obj.employee.department), PDFConfig.ar_text("القسم / Department : ")),
            (PDFConfig.ar_text(self.obj.employee.job_title), PDFConfig.ar_text("المسمى الوظيفي / Job Title : ")),
            (PDFConfig.ar_text(self.obj.employee.job_title), PDFConfig.ar_text("الرقم الوظيفي / Badget Number : ")),
        ]

        story = []
        for i, (label, value) in enumerate(fields):
            text = f"<b>{label}</b> {value}"
            story.append(Paragraph(text, PDFConfig.styles["Arabic"] if i != 2 else PDFConfig.styles["SectionTitle"]))
            story.append(Spacer(1, 6))

        intro_ar = "بموجب هذا النموذج، نُقر نحن الكليات الأولى بأن الموظف المذكور أعلاه قد استلم عهدة العمل التالية:"
        intro_en = (
            "By this form, we at Al-Kulliyat Al-Oula (First Colleges) acknowledge that the "
            "above-mentioned employee has received the following work-related items."
        )

        story.append(Spacer(1, 20))
        story.append(self.make_paragraph(intro_ar))
        story.append(self.make_paragraph(intro_en, PDFConfig.styles["English"]))
        story.append(Spacer(1, 20))

        return story

    def _signature_section(self):
        return [
            Paragraph(PDFConfig.ar_text("اسم الموظف المستلم/ Employee Name: "), PDFConfig.styles["Arabic"]),
            Paragraph(PDFConfig.ar_text("التوقيع / Signature: ___________________"), PDFConfig.styles["Arabic"]),
            Spacer(1, 20),
        ]

    def _declaration_section(self):
        para_ar = (
            "أُقرّ أنا الموقع أدناه بأنني قد استلمت العهدة الموضحة أعلاه بحالة سليمة.\n"
            "وأتعهد بالمحافظة عليها واستخدامها فقط في الأغراض الوظيفية.\n"
            "كما أتحمل كامل المسؤولية عنها، وفي حال ضياعها أو تلفها نتيجة الإهمال أو الاستخدام الخاطئ،\n"
            "أتحمل التبعات القانونية والمالية المترتبة على ذلك، وفق سياسات وأنظمة الشركة."
        )
        para_en = (
            "I, the undersigned, acknowledge that I have received the above-listed items "
            "in good condition.\nI commit to maintaining and using them solely for work purposes.\n"
            "In case of loss or damage due to negligence or misuse, I agree to bear the legal "
            "and financial consequences in accordance with the company's policies."
        )

        return [
            self.make_paragraph("إقرار الموظف:"),
            self.make_paragraph(para_ar),
            Spacer(1, 10),
            Paragraph(para_en, PDFConfig.styles["English"]),
            Spacer(1, 20),
        ]

    def _devices_section(self):
        headers = [
            "الحالة\nCondition",
            "الملحقات\nAccessories",
            "الرقم التسلسلي\nSerial Number",
            "الماركة\nBrand",
            "نوع الجهاز\nItem Type",
            "م.\nNo.",
        ]
        header_row = [self.make_paragraph(h, self.header_style) for h in headers]
        data = [header_row]

        for i, dc in enumerate(self.obj.devices.select_related("device").all()):
            accessories = "\n".join([f"- {a.name_ar}/{a.name_en}" for a in dc.accessories.all()]) or "-"
            item_text = f"{dc.device.name_ar}/{dc.device.name_en} | {dc.device.device_type.name_ar}/{dc.device.device_type.name_en}"

            row = [
                self.make_paragraph(dc.device.get_status_display()),
                self.make_paragraph(accessories),
                self.make_paragraph(dc.device.serial_number or "-"),
                self.make_paragraph(dc.device.brand or "-"),
                self.make_paragraph(item_text),
                Paragraph(str(i + 1), PDFConfig.styles["SectionTitle"]),
            ]
            data.append(row)

        max_col_lengths = [
            max(len(str(r[j].text)) if hasattr(r[j], "text") else 10 for r in data) for j in range(len(headers))
        ]
        total_width = 550
        min_width, max_width = 40, 160
        scale = total_width / sum(max_col_lengths)
        col_widths = [max(min_width, min(int(l * scale), max_width)) for l in max_col_lengths]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2856")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "DejaVu"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("FONTNAME", (0, 1), (-1, -1), "DejaVu"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 1), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F9F9F9"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D3D3D3")),
            ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#1B2856")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ]))
        return [table, Spacer(1, 20)]


# =============== Excel Exporter ===============
class CustodyExcelBuilder:
    def __init__(self, obj):
        self.obj = obj

    def build(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Custody Receipt"

        # ---------- Styles ----------
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
        header_fill = PatternFill("solid", fgColor="1B2856")
        white_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 1
        # ---------- Title ----------
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="نموذج تسليم عهدة / Custody Receipt Form").font = Font(bold=True, size=14)
        ws.cell(row=row, column=1).alignment = center_align
        row += 2

        # ---------- Employee Info ----------
        def add_full_row(value, align=center_align, bold=False):
            nonlocal row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=value)
            cell.alignment = align
            if bold:
                cell.font = Font(bold=True)
            row += 1

        add_full_row(f"التاريخ / Date: {self.obj.custody_date}", right_align)
        add_full_row(f"اسم الموظف / Employee Name: {self.obj.employee.full_name_ar} / {self.obj.employee.full_name_en}", right_align, bold=True)
        add_full_row(f"القسم / Department: {self.obj.employee.department}", right_align)
        add_full_row(f"المسمى الوظيفي / Job Title: {self.obj.employee.job_title}", right_align)
        add_full_row(f"الرقم الوظيفي / Badge Number: {self.obj.employee.job_title or '-'}", right_align)
        row += 1

        # ---------- Intro ----------
        add_full_row("بموجب هذا النموذج، نُقر نحن الكليات الأولى بأن الموظف المذكور أعلاه قد استلم عهدة العمل التالية:", right_align)
        add_full_row("By this form, we at Al-Kulliyat Al-Oula (First Colleges) acknowledge that the employee has received the following items.")
        row += 1

        # ---------- Devices Table ----------
        headers = ["م.\nNo.", "نوع الجهاز\nItem Type", "الماركة\nBrand", "الرقم التسلسلي\nSerial Number", "الملحقات\nAccessories", "الحالة\nCondition"]
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.value = headers[col - 1]
        row += 1

        for i, dc in enumerate(self.obj.devices.select_related("device").all(), start=1):
            accessories = ", ".join([f"{a.name_ar}/{a.name_en}" for a in dc.accessories.all()]) or "-"
            item_text = f"{dc.device.name_ar}/{dc.device.name_en} | {dc.device.device_type.name_ar}/{dc.device.device_type.name_en}"

            values = [
                i,
                item_text,
                dc.device.brand or "-",
                dc.device.serial_number or "-",
                accessories,
                dc.device.get_status_display(),
            ]
            ws.append(values)
            for col in range(1, len(values) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

        row += 1

        # ---------- Declaration ----------
        add_full_row("إقرار الموظف:", right_align, bold=True)
        add_full_row("أُقرّ أنا الموقع أدناه بأنني قد استلمت العهدة الموضحة أعلاه بحالة سليمة. "
                     "وأتعهد بالمحافظة عليها واستخدامها فقط في الأغراض الوظيفية.", right_align)
        add_full_row("I, the undersigned, acknowledge that I have received the above-listed items "
                     "in good condition and commit to maintaining and using them solely for work purposes.")
        row += 1

        # ---------- Signature ----------
        add_full_row("اسم الموظف المستلم / Employee Name:", right_align)
        add_full_row("التوقيع / Signature: ___________________", right_align)

        # ---------- Column widths ----------
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 25

        # Save to BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

# =============== Django View =============== 
class CustodyExportView(PermissionMixin, ExportMixin):
    model = Custody
    filename_prefix = "custody"
    permission_required = "devices.change_custody"

    def get(self, request, *args, **kwargs):
        action = kwargs.get("action")
        obj = self.get_object()

        if action == "pdf":
            builder = CustodyPDFBuilder(obj)
            buffer = builder.build()
            filename = f"{self.filename_prefix}_{obj.pk}.pdf"
            return FileResponse(buffer, as_attachment=True, filename=filename)

        elif action == "excel":
            builder = CustodyExcelBuilder(obj)
            buffer = builder.build()
            filename = f"{self.filename_prefix}_{obj.pk}.xlsx"
            return FileResponse(buffer, as_attachment=True, filename=filename)

        return HttpResponse("Invalid export action", status=400)
